def changePaw(wbN,idC=1,sR=1,eR=1000,npC=2):
    """批量修改工作簿内的学籍号的密码为工作簿内的密码。"""
    try:
        wbN=str(wbN)
        idC=int(idC)
        sR=int(sR)
        eR=int(eR)+1
        npC=int(npC)
    except:
        print("参数错误！")
        return False
    print("Loading…")
    from urllib.request import urlopen
    from urllib.parse import urlencode
    from openpyxl import load_workbook
    wb=load_workbook(wbN)
    ws=wb.active
    print("Ready…change!")
    for i in range(sR,eR):
        if ws.cell(row=i,column=idC).value != None:
            data={}
            url="http://www.lcyzczb.com/cjcx/login_7.asp"
            data["sid"]=ws.cell(row=i,column=idC).value
            data['spassword']=ws.cell(row=i,column=npC).value
            data=urlencode(data)
            dataEc=data.encode()
            urlopen(url,dataEc)
            print(data)
        else:
            break
    print('success!')
if __name__ == '__main__':
    print("以下要求输入的只有工作簿名是必填的，其它若选择不填请直接回车，不要输入内容！")
    wbN=input("请输入存有学籍号的工作簿路径和名（支持相对路径）：")
    print("请确保学籍号在第一个工作表的同一列！")
    inputClo=input("请输入学籍号所在的列（数字，A列为1，B列为2，以此类推）（默认为1）：")
    startRow=input("请输入开始的行（按照工作表行，从1开始，不支持0）（默认为1）：")
    endRow=input("请输入结束的行（按照工作表行，从1开始，不支持0）（默认为1000）：")
    outputClo=input("请输入密码所在的列（数字，A列为1，B列为2，以此类推）（默认为2）：")
    iC=inputClo if inputClo != '' else 1
    sR=startRow if startRow != '' else 1
    eR=endRow if endRow != '' else 1000
    oC=outputClo if outputClo != '' else 2
    changePaw(wbN,iC,sR,eR,oC)
