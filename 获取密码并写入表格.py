def getPaw(wbN,iC=1,sR=1,eR=1000,oC=2):
    """读取工作簿中的学籍号数据，获取密码并输出到同一工作簿。"""
    try:
        wbN=str(wbN)
        iC=int(iC)
        sR=int(sR)
        eR=int(eR)+1
        oC=int(oC)
    except:
        print("参数错误！")
        return False
    print("Loading…")
    from urllib.request import urlopen
    from openpyxl import load_workbook
    wb=load_workbook(wbN)
    ws=wb.active
    print("Ready…Get!")
    for i in range(sR,eR):
        try:
            url="http://www.lcyzczb.com/cjcx/edut_student_7.asp?id="+ws.cell(row=i,column=iC).value if ws.cell(row=i,column=iC).value != None else 'n'
            if url != 'n':
                html=urlopen(url).read().decode('gbk')
                a=html.find('           <td colspan="2"><input name="spassword" type="text" id="spassword" value="')+85 if html.find('           <td colspan="2"><input name="spassword" type="text" id="spassword" value="') > -1 else -1
                b=html.find('" /></td>',a)
                paw=html[a:b] if a > -1 else '<#{"error":"没有此账号"}#>'
            else:
                break
        except:
            paw='<#{"error":"无法解码"}#>'
        ws.cell(row=i,column=oC).value=paw
        print(paw)
    print('Write success,saving……')
    s=wbN.find("\\") if wbN.find("\\") > -1 else 0
    wb.save(wbN[s:wbN.find(".")]+"_Return.xlsx")
    print('success!')

if __name__ == '__main__':
    print("以下要求输入的只有工作簿名是必填的，其它若选择不填请直接回车，不要输入内容！")
    wbN=input("请输入存有学籍号的工作簿路径和名（支持相对路径）：")
    print("请确保学籍号在第一个工作表的同一列！")
    inputClo=input("请输入学籍号所在的列（数字，A列为1，B列为2，以此类推）（默认为1）：")
    startRow=input("请输入学籍号开始的行（按照工作表行，从1开始，不支持0）（默认为1）：")
    endRow=input("请输入学籍号结束的行（按照工作表行，从1开始，不支持0）（默认为1000）：")
    outputClo=input("请输入需要输出到的列（数字，A列为1，B列为2，以此类推）（默认为2）：")
    iC=inputClo if inputClo != '' else 1
    sR=startRow if startRow != '' else 1
    eR=endRow if endRow != '' else 1000
    oC=outputClo if outputClo != '' else 2
    getPaw(wbN,iC,sR,eR,oC)
